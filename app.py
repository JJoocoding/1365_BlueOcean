# ==========================================
# Part 1 — Imports, 기본 설정, 스타일, 공통 유틸
# ==========================================

import itertools
import json
import os
from datetime import datetime

import altair as alt
import numpy as np
import pandas as pd
import requests
import streamlit as st
import xmltodict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


# -------------------------------------------------
# 기본 설정 & SERVICE_KEY 로드
# -------------------------------------------------
st.set_page_config(page_title="1365 사정율 분석기", layout="wide")

try:
    SERVICE_KEY = st.secrets["SERVICE_KEY"]
except Exception:
    SERVICE_KEY = ""


# -------------------------------------------------
# 공통 Request Header
# -------------------------------------------------
def get_headers():
    return {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}


# -------------------------------------------------
# API Header 파싱(JSON/XML)
# -------------------------------------------------
def parse_api_header_from_json(data):
    try:
        response = data.get("response", {})
        header = response.get("header", {})
        return header.get("resultCode"), header.get("resultMsg")
    except Exception:
        return None, None


def parse_api_header_from_xml(data):
    try:
        response = data.get("response", {})
        header = response.get("header", {})
        return header.get("resultCode"), header.get("resultMsg")
    except Exception:
        return None, None


# -------------------------------------------------
# JSON API 호출
# -------------------------------------------------
def fetch_json(url, desc, api_warnings, timeout=10):
    try:
        res = requests.get(url, headers=get_headers(), timeout=timeout)
        res.raise_for_status()
    except Exception as e:
        api_warnings.append(f"[HTTP 오류] {desc}: {e}")
        return None

    try:
        data = json.loads(res.text)
    except Exception as e:
        api_warnings.append(f"[파싱 오류] {desc}: {e}")
        return None

    code, msg = parse_api_header_from_json(data)
    if code is not None and code != "00":
        api_warnings.append(f"[API 오류] {desc}: resultCode={code}, msg={msg}")
        return None

    return data


# -------------------------------------------------
# XML API 호출
# -------------------------------------------------
def fetch_xml(url, desc, api_warnings, timeout=10):
    try:
        res = requests.get(url, headers=get_headers(), timeout=timeout)
        res.raise_for_status()
    except Exception as e:
        api_warnings.append(f"[HTTP 오류] {desc}: {e}")
        return None

    try:
        data = xmltodict.parse(res.text)
    except Exception as e:
        api_warnings.append(f"[파싱 오류] {desc}: {e}")
        return None

    code, msg = parse_api_header_from_xml(data)
    if code is not None and code != "00":
        api_warnings.append(f"[API 오류] {desc}: resultCode={code}, msg={msg}")
        return None

    return data


# -------------------------------------------------
# API 응답 items 추출 유틸
# -------------------------------------------------
def safe_get_items(json_data):
    try:
        response = json_data.get("response", {})
        body = response.get("body", {})
        items = body.get("items")

        if not items:
            return []

        # 리스트 → 그대로 반환
        if isinstance(items, list):
            return items

        # dict 형태
        if isinstance(items, dict):
            item = items.get("item")
            if not item:
                return []
            if isinstance(item, list):
                return item
            if isinstance(item, dict):
                return [item]

        return []
    except Exception:
        return []


# -------------------------------------------------
# Streamlit UI 스타일 정의
# -------------------------------------------------
st.markdown(
    """
<style>
html, body, [data-testid="stAppViewContainer"] {
    background: linear-gradient(135deg, #1e1e2f 0%, #2f2f46 50%, #191926 100%);
    color: #fff !important;
}
.fade-in {
    opacity: 0;
    animation: fadeIn 1.2s forwards;
}
@keyframes fadeIn {
    to { opacity: 1; }
}
.metric-card {
    background: rgba(255,255,255,0.1);
    padding: 18px;
    border-radius: 15px;
    backdrop-filter: blur(8px);
    border: 1px solid rgba(255,255,255,0.2);
    text-align: center;
    transition: 0.3s;
}
.metric-card:hover {
    transform: translateY(-4px);
}
.glow-box {
    background: rgba(255,240,200,0.15);
    border: 1px solid #ffdd9c;
    border-radius: 15px;
    padding: 20px;
    animation: glow 3s infinite ease-in-out;
}
@keyframes glow {
    0% { box-shadow: 0 0 10px #ffdd9c55; }
    50% { box-shadow: 0 0 20px #ffdd9c; }
    100% { box-shadow: 0 0 10px #ffdd9c55; }
}
</style>
""",
    unsafe_allow_html=True,
)


# -------------------------------------------------
# 제목 표시
# -------------------------------------------------
st.markdown(
    """
<h1 class="fade-in" style="text-align:center;
 font-size:40px; font-weight:900;
 background: linear-gradient(90deg,#ffddaa,#ffd087,#ffb067);
 -webkit-background-clip:text; color:transparent;">
🏗 1365 사정율 분석기<br>(핫존 + 블루오션 + 추천 사정률)
</h1>
""",
    unsafe_allow_html=True,
)

# ==========================================
# Part 2 — A값, 집행관, 핫존, 블루오션, 승률분석, 방향성 예측
# ==========================================


# -------------------------------------------------
# A값 조회
# -------------------------------------------------
def get_a_value(gongo_no: str, api_warnings: list) -> float:
    """A값(안전관리비 등) 조회"""
    try:
        url = (
            "http://apis.data.go.kr/1230000/ad/BidPublicInfoService/"
            "getBidPblancListInfoCnstwkBsisAmount"
            f"?inqryDiv=2&bidNtceNo={gongo_no}&pageNo=1&numOfRows=10&type=json&ServiceKey={SERVICE_KEY}"
        )
        data = fetch_json(url, f"A값 조회({gongo_no})", api_warnings)
        if data is None:
            return 0.0

        items = safe_get_items(data)
        if not items:
            return 0.0

        df = pd.DataFrame(items)
        cost_cols = [
            "sftyMngcst",
            "sftyChckMngcst",
            "rtrfundNon",
            "mrfnHealthInsrprm",
            "npnInsrprm",
            "odsnLngtrmrcprInsrprm",
            "qltyMngcst",
        ]
        valid = [c for c in cost_cols if c in df.columns]
        if not valid:
            return 0.0

        return (
            df[valid]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0.0)
            .sum(axis=1)
            .iloc[0]
        )
    except Exception:
        return 0.0


# -------------------------------------------------
# 집행관 이름 조회
# -------------------------------------------------
def get_officer_name_final(gongo_no: str, api_warnings: list) -> str:
    url = (
        "http://apis.data.go.kr/1230000/ad/BidPublicInfoService/"
        f"getBidPblancListInfoCnstwk?inqryDiv=2&bidNtceNo={gongo_no}"
        f"&pageNo=1&numOfRows=1&type=json&ServiceKey={SERVICE_KEY}"
    )
    data = fetch_json(url, f"집행관 조회({gongo_no})", api_warnings)
    if data is None:
        return "확인불가"

    items = safe_get_items(data)
    if not items:
        return "확인불가"

    item = items[0]
    for key in ["exctvNm", "chrgrNm", "ntceChrgrNm"]:
        if key in item and str(item[key]).strip():
            return str(item[key]).strip()

    return "확인불가"


# ---------------------------------------------------------
# 🔮 다음 사정률이 100보다 클지/작을지 방향성 예측
# ---------------------------------------------------------
def predict_direction_next(winner_rates):
    """
    winner_rates: 과거 1순위 사정율 리스트
    return: {up_prob: %, down_prob: %, result_text: str}
    """

    rates = [r for r in winner_rates if r > 0]
    n = len(rates)
    if n < 3:
        return {
            "up_prob": None,
            "down_prob": None,
            "result_text": "데이터 부족(3건 미만)"
        }

    # 1) 최근 N건에서 100 초과 비율
    over100 = sum(r > 100 for r in rates[-10:])
    base_prob = over100 / min(10, n)

    # 2) 변화량(sign) 기반 가중치
    diffs = [rates[i] - rates[i-1] for i in range(1, n)]
    pos = sum(d > 0 for d in diffs[-10:])     # 상승 횟수
    neg = sum(d < 0 for d in diffs[-10:])     # 하락 횟수
    if pos + neg > 0:
        trend_prob = pos / (pos + neg)
    else:
        trend_prob = 0.5

    # 3) 전환점 패턴 보조
    signs = [(r > 100) for r in rates]
    turns = sum(signs[i] != signs[i-1] for i in range(1, len(signs)))
    turn_factor = max(0.7, 1 - (turns * 0.05))

    # 종합 확률
    up_prob = (base_prob * 0.5 + trend_prob * 0.5) * turn_factor
    down_prob = 1 - up_prob

    return {
        "up_prob": round(up_prob * 100, 1),
        "down_prob": round(down_prob * 100, 1),
        "result_text": "상승(100 초과)" if up_prob > 0.5 else "하락(100 미만)"
    }


# ---------------------------------------------------------
# 🟠 핫존 탐색
# ---------------------------------------------------------
def find_hot_zone(actual_rates, window=0.3, step=0.05):
    if not actual_rates:
        return None, None, 0

    rates_sorted = sorted(actual_rates)
    min_r, max_r = min(rates_sorted), max(rates_sorted)

    best_start, best_end, best_count = None, None, -1
    start = min_r

    while start <= max_r:
        end = start + window
        count = sum(start <= r <= end for r in rates_sorted)
        if count > best_count:
            best_count = count
            best_start, best_end = start, end
        start += step

    return best_start, best_end, best_count

# ---------------------------------------------------------
# 🔵 블루오션 v3 (업그레이드: 이동평균 적용 버전)
# ---------------------------------------------------------
def find_blue_ocean_v3(theoretical_rates, bidder_rates, start, end, bin_width=0.0005):
    """
    - theoretical_rates: 1365 조합 사정률 리스트
    - bidder_rates: 실제 업체 사정률 리스트
    - start, end: 분석 구간
    - 업그레이드 내용: 점수 노이즈를 제거하기 위해 윈도우 크기 3의 이동평균을 적용
    """

    if start is None or end is None:
        return None, None, None

    theo = [r for r in theoretical_rates if start <= r <= end]
    bids = [r for r in bidder_rates if start <= r <= end]

    if len(theo) == 0 or len(bids) == 0:
        return None, None, None

    bins = np.arange(start, end + bin_width, bin_width)
    if len(bins) < 2:
        bins = np.array([start, end])

    theo_counts, _ = np.histogram(theo, bins=bins)
    bid_counts, bin_edges = np.histogram(bids, bins=bins)

    if theo_counts.sum() == 0:
        return None, None, None

    theo_norm = theo_counts / theo_counts.sum()
    max_theo = theo_norm.max()
    if max_theo <= 0:
        return None, None, None

    rows = []
    for i in range(len(bin_edges) - 1):
        s = bin_edges[i]
        e = bin_edges[i + 1]
        c = (s + e) / 2

        tcount = theo_counts[i]
        bcount = bid_counts[i]

        demand = theo_norm[i] / max_theo if tcount > 0 else 0
        supply_inv = 1.0 / (bcount + 1.0)
        score = demand * supply_inv

        rows.append({
            "center": c,
            "score": score,
            "theo_count": int(tcount),
            "bid_count": int(bcount),
            "start": s,
            "end": e
        })

    if not rows:
        return None, None, None

    df_blue = pd.DataFrame(rows).sort_values("center").reset_index(drop=True)

    # --- [알고리즘 업그레이드 포인트: Rolling Average 적용] ---
    # 윈도우 3으로 이동평균( smoothed_score )을 구하여 군집된 높은 점수 구역 탐지
    df_blue['smoothed_score'] = df_blue['score'].rolling(window=3, center=True, min_periods=1).mean()
    
    # 이동평균 점수가 가장 높은 행을 최적의 지점으로 선택
    best_idx = df_blue['smoothed_score'].idxmax()
    best_row = df_blue.loc[best_idx]
    
    best_range = (best_row['start'], best_row['end'])
    best_center = best_row['center']
    # ------------------------------------------------------

    return df_blue, best_range, best_center


# ---------------------------------------------------------
# 📊 승률 패널 보조 분석
# ---------------------------------------------------------
def analyze_winrate_distribution(winner_rates, hot_start, hot_end):
    if not winner_rates or hot_start is None or hot_end is None:
        return None

    width = (hot_end - hot_start) / 3
    bins = {
        "Lower": (hot_start, hot_start + width),
        "Mid": (hot_start + width, hot_start + 2 * width),
        "Upper": (hot_start + 2 * width, hot_end),
    }

    total = len(winner_rates)
    results = []

    for name, (s, e) in bins.items():
        wins = sum(s <= r <= e for r in winner_rates)
        win_rate = wins / total if total > 0 else 0
        results.append({
            "구간": name,
            "시도(공고수)": total,
            "승리횟수": wins,
            "승률": round(win_rate * 100, 2),
        })

    return results


# ---------------------------------------------------------
# 승률 패널 표시
# ---------------------------------------------------------
def render_winrate_panel(winner_rates, hot_start, hot_end):
    results = analyze_winrate_distribution(winner_rates, hot_start, hot_end)

    if results is None:
        st.info("승률 분석을 위한 데이터가 부족합니다.")
        return

    N = len(winner_rates)

    if N < 10:
        trust, color = "⚠ 표본 부족 (참고용)", "#FF6B6B"
    elif N < 30:
        trust, color = "🔎 중간 신뢰도", "#FFD93D"
    else:
        trust, color = "🔵 신뢰도 높음", "#4CAF50"

    st.markdown(
        f"""
        <div style="padding:12px;
 border-radius:10px; 
             border:1px solid {color}; margin-top:15px;">
            <h3 style="color:{color};
 margin:0;">📊 승률 분석 패널</h3>
            <p style="color:#ccc;
 margin-top:4px;">핫존 내 승률 비교</p>
            <b style="color:{color};">{trust}</b>
        </div>
        """,
        unsafe_allow_html=True,
    )

    df = pd.DataFrame(results)
    st.dataframe(df, use_container_width=True)

    chart = (
        alt.Chart(df)
        .mark_bar()
        .encode(
            x=alt.X("구간:N", title="핫존 구간 (3분할)"),
            y=alt.Y("승률:Q", title="승률 (%)"),
            tooltip=["구간", "승률", "승리횟수", "시도(공고수)"],
            color=alt.Color("구간:N")
        )
        .properties(title="⚡ 핫존 내 구간별 승률 비교")
    )
    st.altair_chart(chart, use_container_width=True)

# ==========================================
# Part 3 — 공고 분석 + 전체 실행 + UI
# ==========================================

# -------------------------------------------------
# 공고 1건 분석
# -------------------------------------------------
def analyze_gongo(gongo_input_str: str, api_warnings: list):
    """
    공고번호 1건 분석
    - df_combined : 1365 조합 + 실제 입찰 업체 사정율
    - info        : dict(오피서/1순위업체/1순위사정율)
    - df_rates    : 1365 조합 사정율 리스트
    - bidder_rates: 해당 공고 모든 업체 사정율 리스트
    """
    try:
        if "-" in gongo_input_str:
            parts = gongo_input_str.split("-")
            gongo_no = parts[0].strip()
            gongo_ord = parts[1].strip()
        else:
            gongo_no = gongo_input_str.strip()
            gongo_ord = "00"

        officer_name = get_officer_name_final(gongo_no, api_warnings)

        # 1) 복수예가 (1365 조합용)
        url1 = (
            "http://apis.data.go.kr/1230000/as/ScsbidInfoService/"
            "getOpengResultListInfoCnstwkPreparPcDetail"
            f"?inqryDiv=2&bidNtceNo={gongo_no}&bidNtceOrd={gongo_ord}"
            f"&pageNo=1&numOfRows=15&type=json&ServiceKey={SERVICE_KEY}"
        )
        data1 = fetch_json(url1, f"복수예가 조회({gongo_no})", api_warnings)
        df_rates = pd.DataFrame()
        base_price = 0.0

        if data1 is not None:
            try:
                items1 = safe_get_items(data1)
                if items1:
                    df1 = pd.json_normalize(items1)
                    if "bssamt" in df1.columns and "bsisPlnprc" in df1.columns:
                        df1 = df1[["bssamt", "bsisPlnprc"]].astype(float)
                        base_price = (
                            df1.iloc[1]["bssamt"] if len(df1) > 1 else df1.iloc[0]["bssamt"]
                        )
                        df1["SA_rate"] = df1["bsisPlnprc"] / df1["bssamt"] * 100

                        if len(df1) >= 4:
                            rates = [
                                np.mean(c)
                                for c in itertools.combinations(df1["SA_rate"], 4)
                            ]
                            df_rates = (
                                pd.DataFrame({"rate": rates})
                                .sort_values("rate")
                                .reset_index(drop=True)
                            )
                            df_rates["조합순번"] = range(1, len(df_rates) + 1)
            except Exception:
                pass

        # 2) 낙찰하한율
        sucs_rate = 0.0
        url2 = (
            "http://apis.data.go.kr/1230000/ad/BidPublicInfoService/"
            "getBidPblancListInfoCnstwk"
            f"?inqryDiv=2&bidNtceNo={gongo_no}&pageNo=1&numOfRows=1&type=json&ServiceKey={SERVICE_KEY}"
        )
        data2 = fetch_json(url2, f"낙찰하한율 조회({gongo_no})", api_warnings)
        if data2 is not None:
            try:
                items2 = safe_get_items(data2)
                if items2 and "sucsfbidLwltRate" in items2[0]:
                    sucs_rate = float(items2[0]["sucsfbidLwltRate"])
            except Exception:
                pass

        # 3) A값
        A_value = get_a_value(gongo_no, api_warnings)

        # 4) 개찰결과 (XML, 전체 업체)
        url4 = (
            "http://apis.data.go.kr/1230000/as/ScsbidInfoService/"
            f"getOpengResultListInfoOpengCompt?serviceKey={SERVICE_KEY}"
            f"&pageNo=1&numOfRows=999&bidNtceNo={gongo_no}"
        )
        data4 = fetch_xml(url4, f"개찰결과 조회({gongo_no})", api_warnings)
        if data4 is None:
            return (
                pd.DataFrame(),
                f"개찰결과 조회 실패({gongo_input_str})",
                None,
                pd.DataFrame(),
                [],
            )

        try:
            items4_raw = data4.get("response", {}).get("body", {}).get("items")
            if isinstance(items4_raw, dict):
                items4 = items4_raw.get("item", [])
            elif isinstance(items4_raw, list):
                items4 = items4_raw
            else:
                items4 = []
            if isinstance(items4, dict):
                items4 = [items4]
            if not isinstance(items4, list):
                items4 = []
        except Exception:
            items4 = []

        df4 = pd.DataFrame(items4)
        top_info = {"winner": "개찰결과 없음", "rate": 0.0, "officer": officer_name}
        bidder_rates_all = []

        if not df4.empty and "bidprcAmt" in df4.columns:
            df4["bidprcAmt"] = pd.to_numeric(df4["bidprcAmt"], errors="coerce")
            df4 = df4.dropna(subset=["bidprcAmt"])

            if not df4.empty:
                top_name = str(df4.iloc[0].get("prcbdrNm", "업체명없음"))

                if sucs_rate > 0 and base_price > 0:
                    numerator = ((df4["bidprcAmt"] - A_value) * 100) / sucs_rate + A_value
                    df4["rate"] = numerator * 100 / base_price
                else:
                    df4["rate"] = 0.0

                bidder_rates_all = df4["rate"].astype(float).tolist()
                top_row = df4.iloc[0]
                top_rate = float(top_row.get("rate", 0.0))

                top_info = {
                    "winner": top_name,
                    "rate": round(top_rate, 5),
                    "officer": officer_name,
                }

                df4_clean = df4.drop_duplicates(subset=["rate"])
                df4_clean = df4_clean[
                    (df4_clean["rate"] >= 90) & (df4_clean["rate"] <= 110)
                ]
                df4_clean = df4_clean[["prcbdrNm", "rate"]].rename(
                    columns={"prcbdrNm": "업체명"}
                )
            else:
                df4_clean = pd.DataFrame()
        else:
            df4_clean = pd.DataFrame()

        # 5) 조합 + 실제 통합 DF
        if not df_rates.empty:
            df_combined = pd.concat(
                [
                    df_rates[["rate"]].assign(업체명=df_rates["조합순번"].astype(str)),
                    df4_clean[["업체명", "rate"]],
                ],
                ignore_index=True,
            )
        else:
            df_combined = df4_clean.copy()

        if not df_combined.empty:
            df_combined = df_combined.sort_values("rate").reset_index(drop=True)
            df_combined["rate"] = df_combined["rate"].round(5)
            df_combined["공고번호"] = gongo_no

        return df_combined, None, top_info, df_rates, bidder_rates_all

    except Exception as e:
        return (
            pd.DataFrame(),
            f"예외 발생 ({gongo_input_str}): {e}",
            None,
            pd.DataFrame(),
            [],
        )


# -------------------------------------------------
# 전체 실행 + 엑셀 저장
# -------------------------------------------------
def process_analysis(target_officer: str, gongo_input: str):
    api_warnings = []

    if not gongo_input.strip():
        stats_empty = {
            "total": 0,
            "filtered": 0,
            "missing": 0,
            "blue_range": "없음",
            "rec_rate": None,
        }
        return (
            "공고번호를 입력해주세요.",
            None,
            None,
            None,
            "분석된 데이터가 없습니다.",
            None,
            None,
            stats_empty,
            None,
            api_warnings,
            [],
            [],
            [],
        )

    if not SERVICE_KEY:
        api_warnings.append("SERVICE_KEY가 설정되어 있지 않습니다 (secrets.toml 확인).")
        stats_empty = {
            "total": 0,
            "filtered": 0,
            "missing": 0,
            "blue_range": "없음",
            "rec_rate": None,
        }
        return (
            "❌ SERVICE_KEY 미설정 (secrets.toml 확인)",
            None,
            None,
            None,
            "SERVICE_KEY 미설정으로 분석 중단",
            None,
            None,
            stats_empty,
            None,
            api_warnings,
            [],
            [],
            [],
        )

    gongo_list = [x.strip() for x in gongo_input.replace(",", "\n").split("\n") if x.strip()]
    target_clean = target_officer.strip()

    logs = []
    results_for_merge = []
    scatter_data = []
    winner_rates = []
    theoretical_rates_all = []
    bidder_rates_all = []

    total = len(gongo_list)
    progress_bar = st.progress(0.0, text="분석 준비 중...")

    for idx, gongo in enumerate(gongo_list, start=1):
        df, err, info, df_rates_raw, bidder_rates = analyze_gongo(gongo, api_warnings)

        if err:
            logs.append(f"❌ {gongo} | 오류: {err}")
        else:
            officer = str(info["officer"]).strip()
            winner = info["winner"]
            w_rate = info["rate"]

            if target_clean:
                if officer != target_clean:
                    logs.append(f"⛔ [제외] {gongo} | 집행관: {officer}")
                else:
                    logs.append(
                        f"✅ [포함] {gongo} | 집행관: {officer} | 1순위: {winner} ({w_rate}%)"
                    )
                    if not df.empty:
                        results_for_merge.append({"gongo": gongo, "df": df, "info": info})
                    if w_rate != 0:
                        winner_rates.append(w_rate)
                        scatter_data.append([w_rate, gongo, winner])
                    if not df_rates_raw.empty:
                        theoretical_rates_all.extend(df_rates_raw["rate"].tolist())
                    if bidder_rates:
                        bidder_rates_all.extend(bidder_rates)
            else:
                logs.append(
                    f"✅ {gongo} | 집행관: {officer} | 1순위: {winner} ({w_rate}%)"
                )
                if not df.empty:
                    results_for_merge.append({"gongo": gongo, "df": df, "info": info})
                if w_rate != 0:
                    winner_rates.append(w_rate)
                    scatter_data.append([w_rate, gongo, winner])
                if not df_rates_raw.empty:
                    theoretical_rates_all.extend(df_rates_raw["rate"].tolist())
                if bidder_rates:
                    bidder_rates_all.extend(bidder_rates)

        progress = idx / total
        progress_bar.progress(progress, text=f"분석 중... ({idx}/{total})")

    if not results_for_merge:
        logs.append("⚠ 유효한 분석 데이터가 없습니다.")
        stats = {
            "total": len(gongo_list),
            "filtered": 0,
            "missing": len(gongo_list),
            "blue_range": "없음",
            "rec_rate": None,
        }
        progress_bar.progress(1.0, text="분석 완료 (결과 없음)")
        return (
            "\n".join(logs),
            None,
            None,
            None,
            "분석된 데이터가 없습니다.",
            None,
            None,
            stats,
            None,
            api_warnings,
            winner_rates,
            theoretical_rates_all,
            bidder_rates_all,
        )

    all_rates = pd.concat([r["df"]["rate"] for r in results_for_merge]).unique()
    merged_df = pd.DataFrame({"rate": all_rates}).sort_values("rate").reset_index(
        drop=True
    )

    col_index_to_winner = {}
    col_index_to_winrate = {}

    for res in results_for_merge:
        df = res["df"]
        info = res["info"]
        gongo_no = df["공고번호"].iloc[0]
        officer = info["officer"]
        winner = info["winner"]
        w_rate = info["rate"]

        col_name = f"{gongo_no}\n[{officer}]\n{winner}"
        sub_df = df[["rate", "업체명"]].rename(columns={"업체명": col_name})
        merged_df = pd.merge(merged_df, sub_df, on="rate", how="outer")
        col_index_to_winner[col_name] = winner
        col_index_to_winrate[col_name] = w_rate

    merged_df = merged_df.sort_values("rate").reset_index(drop=True).fillna("")

    header_row = {"rate": "1순위 사정률(%)"}
    for col in merged_df.columns[1:]:
        wr = col_index_to_winrate.get(col)
        header_row[col] = f"{wr:.4f}" if wr is not None else ""
    merged_display_df = pd.concat(
        [pd.DataFrame([header_row]), merged_df], ignore_index=True
    )

    hot_start, hot_end = None, None
    if winner_rates:
        hot_start, hot_end, _ = find_hot_zone(winner_rates)
        if hot_start is None or hot_end is None:
            hot_start, hot_end = min(winner_rates), max(winner_rates)

    chart_main = None
    if scatter_data:
        chart_df = pd.DataFrame(scatter_data, columns=["rate", "공고번호", "업체명"])
        min_rate = chart_df["rate"].min()
        max_rate = chart_df["rate"].max()

        if hot_start is not None and hot_end is not None:
            def cat(v):
                return "🔥 핫존" if hot_start <= v <= hot_end else "일반"
        else:
            def cat(v):
                return "전체"

        chart_df["구분"] = chart_df["rate"].apply(cat)

        base_chart = alt.Chart(chart_df).encode(
            x=alt.X(
                "rate",
                title="사정율 (%)",
                scale=alt.Scale(
                    domain=[min(min_rate, 98) - 0.2, max(max_rate, 102) + 0.2]
                ),
            ),
            y=alt.Y("공고번호", sort=None, title="공고번호"),
            tooltip=["업체명", "rate", "공고번호", "구분"],
        )

        chart_main = (
            base_chart.mark_circle(size=140)
            .encode(
                color=alt.condition(
                    alt.datum.구분 == "🔥 핫존",
                    alt.value("#FF3B30"),
                    alt.value("#CCCCCC"),
                ),
                tooltip=["업체명", "rate", "공고번호", "구분"],
            )
            .interactive()
        )

    blue_df, best_range, best_center = None, None, None
    if (
        hot_start is not None
        and hot_end is not None
        and theoretical_rates_all
        and bidder_rates_all
    ):
        blue_df, best_range, best_center = find_blue_ocean_v3(
            theoretical_rates_all, bidder_rates_all, hot_start, hot_end, bin_width=0.0005
        )

    chart_gap = None
    blue_desc = ""
    best_range_str = "없음"
    rec_rate = None
    if blue_df is not None and best_range is not None:
        best_range_str = f"{best_range[0]:.3f}% ~ {best_range[1]:.3f}%"
        rec_rate = round(best_range[1], 4)
        blue_plot_df = blue_df.rename(
            columns={"center": "구간중심", "score": "블루오션점수"}
        )
        chart_gap = (
            alt.Chart(blue_plot_df)
            .mark_bar()
            .encode(
                x=alt.X(
                    "구간중심",
                    title="사정율 구간 중심 (%)",
                    scale=alt.Scale(domain=[hot_start, hot_end]),
                ),
                y=alt.Y("블루오션점수", title="블루오션 점수"),
                tooltip=[
                    "구간중심",
                    "블루오션점수",
                    "theo_count",
                    "bid_count",
                ],
            )
            .properties(title="💎 블루오션 탐지 (핫존 내부)")
            .interactive()
        )
        blue_desc = (
            f"- 이 집행관의 핫존(**{hot_start:.3f}% ~ {hot_end:.3f}%**) 안에서\n"
            f" 1365 이론 조합 밀도는 높지만 실제 투찰 업체 수는 상대적으로 적은\n"
            f" **최상위 블루오션 구간**은 👉 **{best_range_str}** 입니다.\n"
        )
        if rec_rate is not None:
            blue_desc += (
                f"- 이 구간의 상단값을 기준으로 **추천 투찰 사정율**은 "
                f"👉 **{rec_rate:.4f}%** 입니다.\n"
            )
    else:
        if not winner_rates:
            blue_desc = (
                "- 현재 유효한 1순위 사정율 데이터가 없어 블루오션을 계산할 수 없습니다.\n"
                " 개찰이 완료된 공고를 더 추가해 보시거나, 일부 공고의 데이터를 다시 확인해 보세요.\n"
            )
        else:
            blue_desc = (
                "- 현재 데이터로는 뚜렷한 블루오션 구간이 통계적으로 드러나지 않았습니다. "
                "공고 수를 더 늘려 보시는 것도 좋습니다.\n"
            )

    total_input = len(gongo_list)
    filtered = len(results_for_merge)
    missing = total_input - filtered
    stats = {
        "total": total_input,
        "filtered": filtered,
        "missing": missing,
        "blue_range": best_range_str,
        "rec_rate": rec_rate,
    }

    if winner_rates and hot_start is not None and hot_end is not None:
        hotzone_text = (
            f"- 실제 1순위 사정율이 가장 많이 몰린 구간(핫존)은 \n"
            f" 👉 **{hot_start:.3f}% ~ {hot_end:.3f}%** 입니다.\n"
        )
    else:
        hotzone_text = (
            "- 유효한 1순위 사정율이 부족하여 핫존을 계산할 수 없습니다.\n"
            " 개찰이 완료된 공고를 더 추가해 주세요.\n"
        )

    analysis_text = f"""
- 입력 공고 수: **{total_input}건**
- 집행관 필터 통과 공고 수: **{filtered}건**
- 분석에 사용된 1순위 사정율 개수: **{len(winner_rates)}개**

### 🔥 집행관 핫존
{hotzone_text}

### 💎 블루오션 해석
{blue_desc}
"""

    excel_filename = f"사정율분석_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "통합분석"
    for r in dataframe_to_rows(merged_df, index=False, header=True):
        ws.append(r)
    second_row = ["1순위 사정률(%)"]
    for col in merged_df.columns[1:]:
        wr = col_index_to_winrate.get(col)
        second_row.append(f"{wr:.4f}" if wr is not None else "")
    ws.insert_rows(2)
    for col_idx, v in enumerate(second_row, start=1):
        ws.cell(row=2, column=col_idx, value=v)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_align
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = header_align
    fill_winner = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_idx, col_name in enumerate(merged_df.columns, start=1):
        if col_idx == 1: continue
        winner_name = col_index_to_winner.get(col_name)
        if not winner_name: continue
        for row_idx in range(3, ws.max_row + 1):
            if ws.cell(row=row_idx, column=col_idx).value == winner_name:
                ws.cell(row=row_idx, column=col_idx).fill = fill_winner
    if rec_rate is not None:
        highlight_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        lower = rec_rate - 0.0001
        upper = rec_rate + 0.0001
        for row_idx in range(3, ws.max_row + 1):
            rate_value = ws.cell(row=row_idx, column=1).value
            try:
                rate_float = float(rate_value)
                if lower <= rate_float <= upper:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = highlight_fill
            except Exception: pass
    wb.save(excel_filename)
    excel_path = excel_filename
    progress_bar.progress(1.0, text="분석 완료")
    return (
        "\n".join(logs),
        merged_display_df,
        hot_start,
        hot_end,
        analysis_text,
        chart_main,
        chart_gap,
        stats,
        excel_path,
        api_warnings,
        winner_rates,
        theoretical_rates_all,
        bidder_rates_all,
    )

# -------------------------------------------------
# Streamlit UI (디자인 + 승률 패널 + 수동 블루오션)
# -------------------------------------------------
def reset_gongo():
    st.session_state["gongo_text"] = ""


# 간단한 스타일
st.markdown(
    """
<style>
html, body, [data-testid="stAppViewContainer"] {
    background: linear-gradient(135deg, #1e1e2f 0%, #2f2f46 50%, #191926 100%);
    color: #fff !important;
}
.fade-in {
    opacity: 0;
    animation: fadeIn 1.2s forwards;
}
@keyframes fadeIn {
    to { opacity: 1; }
}
.metric-card {
    background: rgba(255,255,255,0.1);
    padding: 18px;
    border-radius: 15px;
    backdrop-filter: blur(8px);
    border: 1px solid rgba(255,255,255,0.2);
    text-align: center;
    transition: 0.3s;
}
.metric-card:hover {
    transform: translateY(-4px);
}
.glow-box {
    background: rgba(255,240,200,0.15);
    border: 1px solid #ffdd9c;
    border-radius: 15px;
    padding: 20px;
    animation: glow 3s infinite ease-in-out;
}
@keyframes glow {
    0% { box-shadow: 0 0 10px #ffdd9c55; }
    50% { box-shadow: 0 0 20px #ffdd9c; }
    100% { box-shadow: 0 0 10px #ffdd9c55; }
}
</style>
""",
    unsafe_allow_html=True,
)



st.markdown("<br>", unsafe_allow_html=True)

target = st.text_input("🎯 타겟 집행관 (선택 사항)", value="")

gongo_input = st.text_area(
    "📄 공고번호 목록 입력",
    height=180,
    key="gongo_text",
    placeholder="예)\nR25BK01074208-000\nR25BK01071774-000\n...",
)

# ----------------------------------------
# 🔧 수동 블루오션 분석 구간 설정 (텍스트 입력 후 float 변환)
# ----------------------------------------
st.markdown("### 🔧 수동 블루오션 분석 구간 설정 (선택사항)")

manual_col1, manual_col2 = st.columns(2)

with manual_col1:
    manual_start_str = st.text_input(
        "수동 구간 시작값 (%)",
        value="",
        placeholder="예: 99.850"
    )

with manual_col2:
    manual_end_str = st.text_input(
        "수동 구간 끝값 (%)",
        value="",
        placeholder="예: 100.020"
    )

manual_start = None
manual_end = None
parse_error = False

try:
    if manual_start_str.strip():
        manual_start = float(manual_start_str.strip())
    if manual_end_str.strip():
        manual_end = float(manual_end_str.strip())
except ValueError:
    parse_error = True
    st.warning("수동 블루오션 구간은 숫자 형식으로 입력해주세요. (예: 99.850)")

btn_col1, btn_col2 = st.columns([1, 1])
with btn_col1:
    run_clicked = st.button("🚀 분석 실행", use_container_width=True)
with btn_col2:
    st.button("🧹 초기화", use_container_width=True, on_click=reset_gongo)

if run_clicked:
    with st.spinner("⏳ 분석 중입니다..."):
        (
            logs,
            merged,
            hot_start,
            hot_end,
            analysis_md,
            chart_main,
            chart_gap,
            stats,
            excel_path,
            api_warnings,
            winner_rates,
            theoretical_rates_all,
            bidder_rates_all,
        ) = process_analysis(target, gongo_input)

    st.session_state["analysis_result"] = {
        "logs": logs,
        "merged": merged,
        "hot_start": hot_start,
        "hot_end": hot_end,
        "analysis_md": analysis_md,
        "chart_main": chart_main,
        "chart_gap": chart_gap,
        "stats": stats,
        "excel_path": excel_path,
        "api_warnings": api_warnings,
        "winner_rates": winner_rates,
        "theoretical_rates_all": theoretical_rates_all,
        "bidder_rates_all": bidder_rates_all,
    }

if "analysis_result" in st.session_state:
    res = st.session_state["analysis_result"]

    logs = res["logs"]
    merged = res["merged"]
    hot_start = res["hot_start"]
    hot_end = res["hot_end"]
    analysis_md = res["analysis_md"]
    chart_main = res["chart_main"]
    chart_gap = res["chart_gap"]
    stats = res["stats"]
    excel_path = res["excel_path"]
    api_warnings = res.get("api_warnings", [])
    winner_rates = res.get("winner_rates", [])
    theoretical_rates_all = res.get("theoretical_rates_all", [])
    bidder_rates_all = res.get("bidder_rates_all", [])

    if api_warnings:
        st.warning(
            "⚠ 공공데이터포털 API 경고/오류가 발생했습니다:\n\n"
            + "\n".join(f"- {w}" for w in api_warnings)
        )

    st.markdown("## 📜 로그")
    st.code(logs or "로그 없음", language="text")

    if merged is None or merged.empty:
        st.error("⚠ 유효한 분석 데이터가 없습니다.")
    else:
        st.markdown("## 🔍 핵심 요약")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            if hot_start is not None:
                c1.markdown(
                    f"<div class='metric-card'><h3>핫존 시작</h3><h2>{hot_start:.4f}%</h2></div>",
                    unsafe_allow_html=True,
                )
            else:
                c1.markdown(
                    "<div class='metric-card'><h3>핫존 시작</h3><h2>N/A</h2></div>",
                    unsafe_allow_html=True,
                )
        with c2:
            if hot_end is not None:
                c2.markdown(
                    f"<div class='metric-card'><h3>핫존 끝</h3><h2>{hot_end:.4f}%</h2></div>",
                    unsafe_allow_html=True,
                )
            else:
                c2.markdown(
                    "<div class='metric-card'><h3>핫존 끝</h3><h2>N/A</h2></div>",
                    unsafe_allow_html=True,
                )
        with c3:
            c3.markdown(
                f"<div class='metric-card'><h3>분석 공고</h3><h2>{stats.get('filtered',0)}</h2></div>",
                unsafe_allow_html=True,
            )
        with c4:
            c4.markdown(
                f"<div class='metric-card'><h3>누락 공고</h3><h2>{stats.get('missing',0)}</h2></div>",
                unsafe_allow_html=True,
            )

        # 추천 사정률
        rec = stats.get("rec_rate")
        st.markdown("## 🔥 추천 투찰 사정률")
        if rec is not None:
            st.markdown(
                f"""
<div class='glow-box'>
    <h2 style='color:#ffcc66;'>🔥 {rec:.4f}%</h2>
    <p style='font-size:14px;'>핫존 + 블루오션 통계 기반 추천값</p>
</div>
""",
                unsafe_allow_html=True,
            )
        else:
            st.info("추천 사정률을 계산할 수 있는 블루오션 구간이 없습니다.")

        # ---------------------------------------------------------
        # 🔮 다음 사정률 방향성 예측
        # ---------------------------------------------------------
        st.markdown("## 🔮 다음 사정률 방향성 예측")
        pred = predict_direction_next(winner_rates)

        if pred["up_prob"] is None:
            st.info("방향성 예측을 위한 데이터가 부족합니다. (3건 이상 필요)")
        else:
            result_color = "#4CAF50" if pred["result_text"].startswith("상승") else "#FF5252"

            st.markdown(
                f"""
                <div style="padding:18px; border-radius:12px;
                     background:rgba(255,255,255,0.08); border:1px solid {result_color};">
                    <h3 style="color:{result_color}; margin-bottom:6px;">🔮 예측 결과</h3>
                    <p style="font-size:20px; font-weight:700; color:{result_color};">
                        다음 사정률은 <b>{pred['result_text']}</b> 가능성이 높습니다.
                    </p>
                    <p style="color:#ddd; margin-top:10px;">
                        📈 100 초과 확률: <b>{pred['up_prob']}%</b><br>
                        📉 100 미만 확률: <b>{pred['down_prob']}%</b>
                    </p>
                </div>
                """,
                unsafe_allow_html=True,
            )

        # 📊 승률 분석 패널
        st.markdown("## 📊 승률 분석 (참고용 보조지표)")
        render_winrate_panel(winner_rates, hot_start, hot_end)

        # 🎯 종합 분석 리포트
        st.markdown("## 🎯 종합 분석 리포트")
        st.markdown(analysis_md)

        # 📈 1순위 사정율 분포
        if chart_main is not None:
            st.markdown("## 📈 1순위 사정율 분포 (핫존 강조)")
            st.altair_chart(chart_main, use_container_width=True)

        # 💎 기본 블루오션 그래프 (핫존 기준)
        if chart_gap is not None:
            st.markdown("## 💎 블루오션 점수 분포 (핫존 내부)")
            st.altair_chart(chart_gap, use_container_width=True)

        # ----------------------------------------
        # 🎯 사용자 지정 블루오션 구간 시각화
        # ----------------------------------------
        if manual_start is not None and manual_end is not None and not parse_error:
            # 시작/끝 뒤집혔으면 자동 보정
            ms, me = manual_start, manual_end
            if ms > me:
                ms, me = me, ms

            st.markdown(
                f"### 🎯 사용자 지정 블루오션 점수 분포 ({ms:.3f}% ~ {me:.3f}%)"
            )

            if not theoretical_rates_all or not bidder_rates_all:
                st.info("수동 블루오션을 계산할 수 있는 이론/업체 데이터가 부족합니다.")
            else:
                manual_blue_df, manual_best_range, manual_best_center = find_blue_ocean_v3(
                    theoretical_rates_all,
                    bidder_rates_all,
                    ms,
                    me,
                    bin_width=0.0005,
                )

                if manual_blue_df is None or manual_best_range is None:
                    st.info("해당 수동 구간에서 통계적으로 의미있는 블루오션 패턴이 보이지 않습니다.")
                else:
                    manual_plot_df = manual_blue_df.rename(
                        columns={"center": "구간중심", "score": "블루오션점수"}
                    )
                    manual_chart = (
                        alt.Chart(manual_plot_df)
                        .mark_bar()
                        .encode(
                            x=alt.X("구간중심", title="사정율 구간 중심 (%)"),
                            y=alt.Y("블루오션점수", title="블루오션 점수"),
                            tooltip=[
                                "구간중심",
                                "블루오션점수",
                                "theo_count",
                                "bid_count",
                            ],
                        )
                        .properties(
                            title=f"💎 사용자 지정 블루오션 탐지 ({ms:.3f}% ~ {me:.3f}%)"
                        )
                        .interactive()
                    )
                    st.altair_chart(manual_chart, use_container_width=True)

        # 📑 통합 테이블
        st.markdown("## 📑 통합 사정율 비교 테이블")
        st.dataframe(merged, use_container_width=True)

        # 📥 엑셀 다운로드
        if excel_path and os.path.exists(excel_path):
            with open(excel_path, "rb") as f:
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=f,
                    file_name=os.path.basename(excel_path),
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )
